from __future__ import annotations

from collections import defaultdict
import sys
from datetime import datetime, date
from pathlib import Path
import shutil
from typing import Dict, List, Optional, Tuple

import pandas as pd
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font

BASE_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))

MEAL_ORDER = ["Breakfast", "Lunch", "Hi-tea", "Dinner"]

MEAL_LABELS_EN = {
    "Breakfast": "Breakfast",
    "Lunch": "Lunch",
    "Hi-tea": "Hi-tea",
    "Dinner": "Dinner",
}

MEAL_LABELS_HI = {
    "Breakfast": "नाश्ता",
    "Lunch": "दोपहर का भोजन",
    "Hi-tea": "हाई टी",
    "Dinner": "रात्रि भोजन",
}

CATEGORY_LABELS_HI = {
    "live": "लाइव",
    "starter": "स्टार्टर",
    "main course": "मुख्य व्यंजन",
    "maincourse": "मुख्य व्यंजन",
    "bread": "ब्रेड",
    "raita": "रायता",
    "sides": "साइड्स",
    "salad": "सलाद",
    "condiments": "मसाले",
    "counter": "काउंटर",
    "snacks": "स्नैक्स",
    "beverages": "पेय",
    "dessert": "मिठाई",
    "static": "स्टैटिक",
    "live counter": "लाइव काउंटर",
    "soup": "सूप",
}

LABELS_EN = {
    "client_name": "Client Name",
    "venue": "Venue",
    "total_pax": "Total Pax",
    "start_date": "Start Date",
    "end_date": "End Date",
    "caterer": "Caterer",
    "phone": "Phone",
    "dinner_planner": "Planner",
    "count": "Count",
    "generated_on": "Generated on",
    "brand": "Pushp Events",
}

LABELS_HI = {
    "client_name": "क्लाइंट नाम",
    "venue": "स्थान",
    "total_pax": "कुल पैक्स",
    "start_date": "प्रारंभ तिथि",
    "end_date": "समाप्ति तिथि",
    "caterer": "कैटरर",
    "phone": "फोन",
    "dinner_planner": "प्लानर",
    "count": "गणना",
    "generated_on": "तैयार",
    "brand": "Pushp Events",
}

MONTHS_HI = {
    1: "जनवरी",
    2: "फरवरी",
    3: "मार्च",
    4: "अप्रैल",
    5: "मई",
    6: "जून",
    7: "जुलाई",
    8: "अगस्त",
    9: "सितंबर",
    10: "अक्टूबर",
    11: "नवंबर",
    12: "दिसंबर",
}


# ---------------- HELPERS ----------------

def clean(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


def normalize_key(key: str) -> str:
    key = clean(key).lower()
    key = key.replace("&", "and")
    key = key.replace("-", " ").replace("_", " ")
    key = " ".join(key.split())
    return key


EVENT_KEY_ALIASES = {
    "event_name": ["event name", "event", "event title"],
    "event_name_hi": ["event name hi", "event_name_hi", "event title hi", "event name hindi"],
    "client_name": ["client name", "client", "client_name"],
    "client_name_hi": ["client name hi", "client_name_hi", "client name hindi"],
    "venue": ["venue", "event venue", "location"],
    "venue_hi": ["venue hi", "venue_hindi", "venue_hi", "location hi"],
    "start_date": ["start date", "start", "start date (dd/mm/yyyy)"],
    "end_date": ["end date", "end", "end date (dd/mm/yyyy)"],
    "total_pax": ["total pax", "pax", "total pax size", "total guests"],
    "caterer_name": ["caterer name", "cater's name", "cater name", "caterer"],
    "caterer_phone": [
        "caterer phone",
        "caterer number",
        "cater's number",
        "phone",
        "contact phone",
        "contact_phone",
    ],
    "logo_path": ["logo", "logo path", "logo file", "caterer logo", "cater's logo"],
    "planner_name": ["planner name", "event planner", "dinner planner name", "planner"],
    "planner_name_hi": ["planner name hi", "planner_name_hi", "planner name hindi"],
    "caterer_name_hi": ["caterer name hi", "caterer_name_hi", "caterer name hindi"],
    "notes_title": ["notes title", "note title", "extras title", "notes heading", "extras heading"],
    "notes_title_hi": ["notes title hi", "notes_title_hi", "extras title hi", "notes heading hindi"],
    "notes_text": ["notes", "note", "extras", "notes text", "notes_text", "extras text"],
    "notes_text_hi": ["notes_text_hi", "notes text hi", "extras text hi", "notes hindi", "extras hindi"],
}


def normalize_event_info(raw: Dict[str, object]) -> Dict[str, object]:
    normalized = {}
    reverse_lookup = {}

    for canonical, aliases in EVENT_KEY_ALIASES.items():
        for alias in aliases:
            reverse_lookup[normalize_key(alias)] = canonical

    for key, value in raw.items():
        canonical = reverse_lookup.get(normalize_key(key), normalize_key(key))
        normalized[canonical] = value

    return normalized


def parse_date(value: object) -> Optional[date]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    # Try explicit formats before relying on pandas inference.
    # DD/MM/YYYY and DD-MM-YYYY are the expected input formats.
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        parsed = pd.to_datetime(s, errors="coerce", dayfirst=True)
        if pd.isna(parsed):
            return None
        return parsed.date()
    except Exception:
        return None


def format_date(value: Optional[date]) -> str:
    if not value:
        return ""
    return value.strftime("%d %B %Y")


def format_date_localized(value: Optional[date], lang: str) -> str:
    if not value:
        return ""
    if lang == "hi":
        month = MONTHS_HI.get(value.month, value.strftime("%B"))
        return f"{value.day:02d} {month} {value.year}"
    return value.strftime("%d %B %Y")


def normalize_meal(meal: str) -> str:
    raw = normalize_key(meal)
    if not raw:
        return ""
    if raw in {"breakfast", "bf"}:
        return "Breakfast"
    if raw in {"lunch"}:
        return "Lunch"
    if raw in {"hi tea", "hi-tea", "hitea", "high tea", "high-tea"}:
        return "Hi-tea"
    if raw in {"dinner"}:
        return "Dinner"
    return clean(meal)


def parse_note_lines(raw_text: object) -> List[str]:
    text = clean(raw_text)
    if not text:
        return []

    lines = []
    for line in text.splitlines():
        item = line.strip()
        if not item:
            continue
        # Normalize common bullet prefixes from manual entry.
        item = item.lstrip("-*•").strip()
        if not item:
            continue
        for part in item.split(";"):
            chunk = part.strip()
            if chunk:
                lines.append(chunk)

    return lines


# ---------------- EXCEL IO ----------------

def _build_date_meal_exprs() -> Tuple[str, str]:
    """Return (start_expr, formula_total) for Excel date/meal auto-fill formulas."""
    start_ref = 'INDEX(event_info!B:B, MATCH("start_date", event_info!A:A, 0))'
    end_ref = 'INDEX(event_info!B:B, MATCH("end_date", event_info!A:A, 0))'
    start_text = f'IFERROR(TEXT({start_ref},"dd/mm/yyyy"), {start_ref})'
    end_text = f'IFERROR(TEXT({end_ref},"dd/mm/yyyy"), {end_ref})'
    s1 = f'FIND("/",{start_text})'
    s2 = f'FIND("/",{start_text},{s1}+1)'
    e1 = f'FIND("/",{end_text})'
    e2 = f'FIND("/",{end_text},{e1}+1)'
    start_d = (
        f'DATE(VALUE(RIGHT({start_text},4)),'
        f'VALUE(MID({start_text},{s1}+1,{s2}-{s1}-1)),'
        f'VALUE(LEFT({start_text},{s1}-1)))'
    )
    end_d = (
        f'DATE(VALUE(RIGHT({end_text},4)),'
        f'VALUE(MID({end_text},{e1}+1,{e2}-{e1}-1)),'
        f'VALUE(LEFT({end_text},{e1}-1)))'
    )
    start_expr = f'IF(ISNUMBER({start_ref}),{start_ref},IFERROR({start_d},DATEVALUE({start_text})))'
    end_expr = f'IF(ISNUMBER({end_ref}),{end_ref},IFERROR({end_d},DATEVALUE({end_text})))'
    formula_total = f'({end_expr} - {start_expr} + 1) * 4'
    return start_expr, formula_total


def read_event_info(xlsx_path: Path) -> Dict[str, object]:
    df = pd.read_excel(xlsx_path, sheet_name="event_info")
    if "key" in df.columns and "value" in df.columns:
        raw = {
            clean(k): v
            for k, v in zip(df["key"], df["value"])
            if clean(k)
        }
    else:
        row = df.iloc[0].to_dict() if len(df) else {}
        raw = {clean(k): v for k, v in row.items() if clean(k)}
    return normalize_event_info(raw)


def read_menu_data(xlsx_path: Path) -> pd.DataFrame:
    df = pd.read_excel(xlsx_path, sheet_name="menu")
    return df


def get_date_range(
    event_info: Dict[str, object], menu_dates: List[date]
) -> Tuple[Optional[date], Optional[date]]:
    start_date = parse_date(event_info.get("start_date"))
    end_date = parse_date(event_info.get("end_date"))

    if not start_date or not end_date:
        if menu_dates:
            start_date = min(menu_dates)
            end_date = max(menu_dates)

    return start_date, end_date


def ensure_template_exists(template_path: Path, source_excel: Path) -> None:
    if template_path.exists():
        return
    template_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_excel, template_path)


def ensure_meal_counts_sheet(
    xlsx_path: Path,
    date_list: List[date],
    default_pax: object,
    meal_order: List[str] = MEAL_ORDER,
) -> None:
    wb = load_workbook(xlsx_path)
    if "meal_counts" in wb.sheetnames:
        wb.close()
        return

    ws = wb.create_sheet("meal_counts")
    ws.append(["date", "meal", "count"])
    header_fill = PatternFill("solid", fgColor="D9D2C6")
    header_font = Font(bold=True, color="000000")
    for col in ("A1", "B1", "C1"):
        ws[col].fill = header_fill
        ws[col].font = header_font

    date_format = "DD/MM/YYYY"

    # Build formula components (reuses shared helper to avoid duplication)
    start_expr, formula_total = _build_date_meal_exprs()
    formula_date = (
        '=IFERROR(IF(ROW()-1 <= {total}, '
        '{start} + INT((ROW()-2)/4), ""), "")'
    ).format(total=formula_total, start=start_expr)
    formula_meal = (
        '=IFERROR(IF(ROW()-1 <= {total}, '
        'CHOOSE(MOD(ROW()-2,4)+1, "Breakfast", "Lunch", "Hi-tea", "Dinner"), ""), "")'
    ).format(total=formula_total)
    # Count should be entered manually; do not reference total_pax to avoid circular formulas.

    for row_idx in range(2, 1502):
        ws.cell(row_idx, 1, formula_date).number_format = date_format
        ws.cell(row_idx, 2, formula_meal)

    wb.save(xlsx_path)
    wb.close()


def read_meal_counts(
    xlsx_path: Path,
    default_count: Optional[object] = None,
) -> Dict[date, Dict[str, object]]:
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception:
        return {}

    if "meal_counts" not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb["meal_counts"]
    headers = {}
    for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1):
        key = normalize_key(cell.value)
        if key:
            headers[key] = idx

    date_idx = headers.get("date", 1)
    meal_idx = headers.get("meal", 2)
    count_idx = headers.get("count") or headers.get("pax") or 3

    meal_counts: Dict[date, Dict[str, object]] = defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = parse_date(row[date_idx - 1] if len(row) >= date_idx else None)
        meal = normalize_meal(row[meal_idx - 1] if len(row) >= meal_idx else "")
        count_val = row[count_idx - 1] if len(row) >= count_idx else None
        if count_val in (None, "") and default_count not in (None, ""):
            count_val = default_count
        if not d or not meal:
            continue
        meal_counts[d][meal] = count_val

    wb.close()
    return meal_counts


def parse_count(value: object) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        text = str(value).strip()
        if not text:
            return None
        return float(text)
    except Exception:
        return None


def compute_date_counts(
    meal_counts: Dict[date, Dict[str, object]]
) -> Dict[date, Optional[float]]:
    date_counts: Dict[date, Optional[float]] = {}
    for d, meals in meal_counts.items():
        values = [parse_count(v) for v in meals.values()]
        values = [v for v in values if v is not None]
        if values:
            date_counts[d] = sum(values)
        else:
            date_counts[d] = None
    return date_counts


# ---------------- MENU BUILD ----------------

def build_menu_tree(menu_df: pd.DataFrame) -> Dict[date, Dict[str, Dict[str, List[str]]]]:
    menu_tree: Dict[date, Dict[str, Dict[str, List[str]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(list))
    )

    for _, row in menu_df.iterrows():
        row_date = parse_date(row.get("date"))
        meal = normalize_meal(row.get("meal", ""))
        category = clean(row.get("category", ""))
        item = clean(row.get("item", ""))

        if not row_date or not meal or not item:
            continue

        if not category:
            category = "Menu"

        menu_tree[row_date][meal][category].append(item)

    return menu_tree


def build_date_pages(
    date_list: List[date],
    menu_tree: Dict[date, Dict[str, Dict[str, List[str]]]],
    meal_counts: Dict[date, Dict[str, object]],
    date_counts: Dict[date, Optional[float]],
    meal_label_map: Dict[str, str],
    date_formatter,
    meal_order: List[str] = MEAL_ORDER,
) -> List[Dict[str, object]]:
    pages: List[Dict[str, object]] = []

    for d in date_list:
        meals = []
        for meal in meal_order:
            categories = menu_tree.get(d, {}).get(meal, {})
            pax = meal_counts.get(d, {}).get(meal, "")
            if categories or pax or meal in menu_tree.get(d, {}):
                meals.append(
                    {
                        "name": meal,
                        "display_name": meal_label_map.get(meal, meal),
                        "pax": clean(pax),
                        "categories": categories,
                    }
                )

        # append any custom meals not in MEAL_ORDER
        for meal, categories in menu_tree.get(d, {}).items():
            if meal not in meal_order:
                pax = meal_counts.get(d, {}).get(meal, "")
                meals.append(
                    {
                        "name": meal,
                        "display_name": meal_label_map.get(meal, meal),
                        "pax": clean(pax),
                        "categories": categories,
                    }
                )

        left_meals: List[Dict[str, object]] = []
        right_meals: List[Dict[str, object]] = []
        for meal in meals:
            name = meal.get("name", "")
            if name in {"Breakfast", "Hi-tea"}:
                left_meals.append(meal)
            elif name in {"Lunch", "Dinner"}:
                right_meals.append(meal)
            else:
                if len(left_meals) <= len(right_meals):
                    left_meals.append(meal)
                else:
                    right_meals.append(meal)

        pages.append(
            {
                "date": date_formatter(d),
                "count": "" if date_counts.get(d) is None else int(date_counts[d]),
                "meals": meals,
                "left_meals": left_meals,
                "right_meals": right_meals,
            }
        )

    return pages


# ---------------- PDF GENERATION ----------------

def generate_menu_pdf(
    excel_path: Path,
    output_path: Optional[Path] = None,
    template_path: Optional[Path] = None,
    labels: Optional[Dict[str, str]] = None,
    meal_label_map: Optional[Dict[str, str]] = None,
    event_name_override: Optional[str] = None,
    category_labels: Optional[Dict[str, str]] = None,
    lang: str = "en",
) -> Path:
    excel_path = Path(excel_path)
    if output_path is None:
        output_path = excel_path.parent / "output" / "Pushp_Events_Menu.pdf"
    output_path = Path(output_path)

    if template_path is None:
        template_path = None

    event_info = read_event_info(excel_path)

    menu_df = read_menu_data(excel_path)
    menu_tree = build_menu_tree(menu_df)

    menu_dates = sorted(menu_tree.keys())
    start_date, end_date = get_date_range(event_info, menu_dates)

    if start_date and end_date:
        date_list = [d.date() for d in pd.date_range(start_date, end_date, freq="D")]
    else:
        date_list = menu_dates

    default_pax = event_info.get("total_pax", "")
    if isinstance(default_pax, float) and pd.isna(default_pax):
        default_pax = ""

    if date_list:
        ensure_meal_counts_sheet(excel_path, date_list, default_pax)

    meal_counts = read_meal_counts(excel_path, None)
    date_counts = compute_date_counts(meal_counts)

    if labels is None:
        labels = LABELS_EN
    if meal_label_map is None:
        meal_label_map = MEAL_LABELS_EN
    if category_labels is None:
        category_labels = {}

    date_formatter = lambda d: format_date_localized(d, lang)

    date_pages = build_date_pages(
        date_list, menu_tree, meal_counts, date_counts, meal_label_map, date_formatter
    )

    logo_path = clean(event_info.get("logo_path")) or "assets/pushp-event-logo.png"
    font_path = "assets/NotoSerifDevanagari-Regular.ttf"

    total_pax_value = clean(event_info.get("total_pax"))
    if not total_pax_value:
        all_counts = []
        for meals in meal_counts.values():
            for v in meals.values():
                pv = parse_count(v)
                if pv is not None:
                    all_counts.append(pv)
        if all_counts:
            total_pax_value = str(int(sum(all_counts)))

    def localized_value(key: str) -> str:
        if lang == "hi":
            hi_key = f"{key}_hi"
            hi_val = clean(event_info.get(hi_key))
            if hi_val:
                return hi_val
        return clean(event_info.get(key))

    event_name_value = localized_value("event_name")
    if event_name_override:
        event_name_value = event_name_override

    event = {
        "event_name": event_name_value,
        "event_name_hi": clean(event_info.get("event_name_hi")),
        "client_name": localized_value("client_name"),
        "venue": localized_value("venue"),
        "start_date": format_date_localized(start_date, lang),
        "end_date": format_date_localized(end_date, lang),
        "total_pax": total_pax_value,
        "caterer_name": localized_value("caterer_name"),
        "caterer_phone": clean(event_info.get("caterer_phone")),
        "planner_name": localized_value("planner_name"),
        "logo_path": logo_path,
    }

    notes_title = localized_value("notes_title")
    notes_text = localized_value("notes_text")
    if not notes_title:
        notes_title = "नोट्स" if lang == "hi" else "Notes"
    notes_items = parse_note_lines(notes_text)

    env = Environment(loader=FileSystemLoader(str(BASE_DIR / "templates")))
    template = env.get_template("menu.html")

    html_out = template.render(
        event=event,
        date_pages=date_pages,
        logo_path=logo_path,
        font_path=font_path,
        generated_on=format_date_localized(datetime.now().date(), lang),
        labels=labels,
        category_labels=category_labels,
        notes_title=notes_title,
        notes_items=notes_items,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)

    from weasyprint import HTML  # lazy — avoids GTK3 load at startup
    HTML(string=html_out, base_url=BASE_DIR.as_uri() + "/").write_pdf(str(output_path))

    if template_path:
        ensure_template_exists(Path(template_path), excel_path)

    return output_path


def safe_filename(name: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in (" ", "-", "_") else " " for ch in name)
    cleaned = "_".join(cleaned.split())
    return cleaned or "Menu"


def generate_menu_pdfs(
    excel_path: Path, output_root: Optional[Path] = None
) -> Tuple[Path, Path]:
    excel_path = Path(excel_path)
    event_info = read_event_info(excel_path)
    event_name = clean(event_info.get("event_name")) or "Menu"
    event_name_hi = clean(event_info.get("event_name_hi")) or event_name

    if output_root is None:
        output_root = Path.home() / "Documents"
    output_root = Path(output_root)
    menu_root = output_root
    if menu_root.name.lower() != "generated-menu":
        menu_root = menu_root / "Generated-menu"

    output_dir = menu_root / excel_path.stem
    output_dir.mkdir(parents=True, exist_ok=True)

    base_en = safe_filename(event_name)
    base_hi = safe_filename(event_name_hi)

    output_en = output_dir / f"{base_en} [en].pdf"
    output_hi = output_dir / f"{base_hi} [hi].pdf"

    generate_menu_pdf(
        excel_path,
        output_path=output_en,
        labels=LABELS_EN,
        meal_label_map=MEAL_LABELS_EN,
        event_name_override=event_name,
        category_labels={},
        lang="en",
    )
    generate_menu_pdf(
        excel_path,
        output_path=output_hi,
        labels=LABELS_HI,
        meal_label_map=MEAL_LABELS_HI,
        event_name_override=event_name_hi,
        category_labels=CATEGORY_LABELS_HI,
        lang="hi",
    )

    return output_en, output_hi


def create_template_excel(path: Path) -> None:
    path = Path(path)
    wb = Workbook()

    header_fill = "D9D2C6"
    header_font = "000000"
    date_format = "DD/MM/YYYY"

    # event_info
    ws_event = wb.active
    ws_event.title = "event_info"
    ws_event.append(["key", "value"])
    ws_event["A1"].fill = PatternFill("solid", fgColor=header_fill)
    ws_event["B1"].fill = PatternFill("solid", fgColor=header_fill)
    ws_event["A1"].font = Font(bold=True, color=header_font)
    ws_event["B1"].font = Font(bold=True, color=header_font)
    keys = [
        "event_name",
        "event_name_hi",
        "client_name",
        "client_name_hi",
        "venue",
        "venue_hi",
        "city",
        "start_date",
        "end_date",
        "branding_name",
        "contact_phone",
        "total_pax",
        "planner_name",
        "planner_name_hi",
        "caterer_name",
        "caterer_name_hi",
        "logo_path",
        "notes_title",
        "notes_title_hi",
        "notes_text",
        "notes_text_hi",
    ]
    for key in keys:
        if key == "total_pax":
            ws_event.append(
                [key, '=IFERROR(IF(SUM(meal_counts!C:C)=0, "", SUM(meal_counts!C:C)), "")']
            )
        elif key == "notes_title":
            ws_event.append([key, "Notes"])
        elif key == "notes_title_hi":
            ws_event.append([key, "नोट्स"])
        else:
            ws_event.append([key, ""])

    # Format start/end date cells as DD/MM/YYYY date cells (not text).
    # Storing as a date serial avoids locale-driven ambiguity — openpyxl reads
    # serials as datetime objects, which parse_date handles with zero ambiguity.
    # For the edge case where Excel leaves the value as text (e.g. day > 12 on
    # US locale), parse_date explicitly tries DD/MM/YYYY and DD-MM-YYYY first.
    for row_idx, key in enumerate(keys, start=2):
        if key in ("start_date", "end_date"):
            cell = ws_event[f"B{row_idx}"]
            cell.number_format = "DD/MM/YYYY"

    # Add data validation to enforce DD/MM/YYYY for start_date and end_date
    try:
        from openpyxl.worksheet.datavalidation import DataValidation

        start_row = keys.index("start_date") + 2
        end_row = keys.index("end_date") + 2
        for row in (start_row, end_row):
            cell_ref = f"B{row}"
            dv = DataValidation(
                type="custom",
                formula1=(
                    f'=OR({cell_ref}="",ISNUMBER({cell_ref}),'
                    f'AND(LEN({cell_ref})>=8,LEN({cell_ref})<=10,'
                    f'ISNUMBER(FIND("/",{cell_ref})),'
                    f'ISNUMBER(FIND("/",{cell_ref},FIND("/",{cell_ref})+1))))'
                ),
                allow_blank=True,
            )
            dv.error = "Please enter a valid date (e.g., 15/04/2026 or 15-04-2026)."
            dv.errorTitle = "Invalid date"
            dv.prompt = "Enter date as DD/MM/YYYY (e.g., 15/04/2026)."
            dv.promptTitle = "Date"
            dv.showErrorMessage = True
            ws_event.add_data_validation(dv)
            dv.add(ws_event[cell_ref])
    except Exception:
        pass

    # menu sheet
    ws_menu = wb.create_sheet("menu")
    ws_menu.append(["date", "meal", "category", "item"])
    for col in ("A1", "B1", "C1", "D1"):
        ws_menu[col].fill = PatternFill("solid", fgColor=header_fill)
        ws_menu[col].font = Font(bold=True, color=header_font)

    start_expr, formula_total = _build_date_meal_exprs()

    for row_idx in range(2, 1502):
        # Use fixed row index in formula so copied cells keep the same date/meal value.
        formula_date_row = (
            '=IFERROR(IF({row}-1 <= {total}, '
            '{start} + INT(({row}-2)/4), ""), "")'
        ).format(row=row_idx, total=formula_total, start=start_expr)
        formula_meal_row = (
            '=IFERROR(IF({row}-1 <= {total}, '
            'CHOOSE(MOD({row}-2,4)+1, "Breakfast", "Lunch", "Hi-tea", "Dinner"), ""), "")'
        ).format(row=row_idx, total=formula_total)
        ws_menu.cell(row_idx, 1, formula_date_row).number_format = date_format
        ws_menu.cell(row_idx, 2, formula_meal_row)

    # meal_counts sheet
    ws_counts = wb.create_sheet("meal_counts")
    ws_counts.append(["date", "meal", "count"])
    for col in ("A1", "B1", "C1"):
        ws_counts[col].fill = PatternFill("solid", fgColor=header_fill)
        ws_counts[col].font = Font(bold=True, color=header_font)
    # Count is manual to avoid circular formulas.

    formula_date_counts = (
        '=IFERROR(IF(ROW()-1 <= {total}, '
        '{start} + INT((ROW()-2)/4), ""), "")'
    ).format(total=formula_total, start=start_expr)
    formula_meal_counts = (
        '=IFERROR(IF(ROW()-1 <= {total}, '
        'CHOOSE(MOD(ROW()-2,4)+1, "Breakfast", "Lunch", "Hi-tea", "Dinner"), ""), "")'
    ).format(total=formula_total)

    for row_idx in range(2, 1502):
        ws_counts.cell(row_idx, 1, formula_date_counts).number_format = date_format
        ws_counts.cell(row_idx, 2, formula_meal_counts)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ---------------- NAME TAGS ----------------

def get_all_menu_items(excel_path: Path) -> List[Dict[str, str]]:
    """Return unique menu items (deduped by name) sorted alphabetically."""
    menu_df = read_menu_data(excel_path)
    seen: set = set()
    items = []
    for _, row in menu_df.iterrows():
        item = clean(row.get("item", ""))
        if not item or item in seen:
            continue
        seen.add(item)
        items.append({
            "item": item,
            "category": clean(row.get("category", "")),
            "meal": normalize_meal(row.get("meal", "")),
        })
    return sorted(items, key=lambda x: x["item"].lower())


def generate_name_tags_pdf(excel_path: Path, output_path: Path) -> Path:
    """Generate a label-sheet PDF with one name tag per unique menu item."""
    import base64
    from weasyprint import HTML  # lazy — avoids GTK3 load at startup

    items = get_all_menu_items(excel_path)
    event_info = read_event_info(excel_path)

    caterer_name = clean(event_info.get("caterer_name", "")) or "PUSHP EVENTS"
    caterer_phone = clean(event_info.get("caterer_phone", ""))

    # Resolve logo: prefer event_info path, then default asset
    logo_data_uri = ""
    logo_path_raw = clean(event_info.get("logo_path", ""))
    candidates = [
        Path(logo_path_raw) if logo_path_raw else None,
        BASE_DIR / "assets" / "pushp-event-logo.png",
        BASE_DIR / "assets" / "logo.png",
    ]
    for candidate in candidates:
        if candidate and candidate.exists():
            mime = "image/png" if candidate.suffix.lower() == ".png" else "image/jpeg"
            logo_data_uri = (
                f"data:{mime};base64,"
                + base64.b64encode(candidate.read_bytes()).decode()
            )
            break

    env = Environment(loader=FileSystemLoader(str(BASE_DIR / "templates")))
    template = env.get_template("nametags.html")
    font_path = "assets/NotoSerifDevanagari-Regular.ttf"
    html_out = template.render(
        items=items,
        font_path=font_path,
        caterer_name=caterer_name,
        caterer_phone=caterer_phone,
        logo_data_uri=logo_data_uri,
    )
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    HTML(string=html_out, base_url=BASE_DIR.as_uri() + "/").write_pdf(str(output_path))
    return output_path


# ---------------- RESET ----------------

def reset_excel(
    excel_path: Path,
    template_path: Optional[Path] = None,
    create_new: bool = True,
) -> Path:
    excel_path = Path(excel_path)
    if create_new:
        reset_path = excel_path.parent / f"{excel_path.stem}_reset.xlsx"
        create_template_excel(reset_path)
        return reset_path

    create_template_excel(excel_path)
    return excel_path
